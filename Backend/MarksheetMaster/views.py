from django.shortcuts import render

# Create your views here.
from django.db.models import Subquery, OuterRef

from rest_framework.response import Response
import json
from rest_framework.generics import GenericAPIView
from MarksheetMaster.models import *
from SchoolMaster.models import *
from SchoolMaster.serializers import *
from Parent_StudentMaster.models import *
from Parent_StudentMaster.serializers import *
from User.models import *
from MarksheetMaster.serializers import *
from rest_framework.authentication import (BaseAuthentication,
                                           get_authorization_header)
from rest_framework import permissions
from User.jwt import userJWTAuthentication
from datetime import datetime
from tablib import Dataset
from openpyxl import load_workbook
import pandas as pd
import os
from django.conf import settings

from rest_framework.parsers import FileUploadParser


class add_exam_type(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        data['isActive'] = True
        data['school_code'] =request.user.school_code
        examtypeexist = ExamType.objects.filter(TypeName=data['TypeName'],isActive= True,school_code=data['school_code']).first()
        if examtypeexist is not None:
            return Response({"data":'',"response": {"n": 0, "msg": "Type Name with marks already exist","status": "failure"}})
        else:
            serializer = ExamTypeSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam Type added successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Exam Type not added ","status": "failure"}})

        
class ExamTypelist(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def get(self,request):
        ExamTypeobjs = ExamType.objects.filter(isActive=True,school_code=request.user.school_code).order_by('-id')
        serializer = ExamTypeSerializer(ExamTypeobjs,many=True)
        return Response({"data":serializer.data,"response": {"n": 1, "msg": "ExamType list found successfully","status": "success"}})
      
       
class ExamTypebyid(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        id = request.data.get('id')
        ExamTypeobj = ExamType.objects.filter(id=id,isActive=True,school_code=request.user.school_code).first()
        if ExamTypeobj is not None:
            serializer = ExamTypeSerializer(ExamTypeobj)
            return Response({"data":serializer.data,"response": {"n": 1, "msg": "ExamType found successfully","status": "success"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "ExamType not found ","status": "failure"}})



class updateExamType(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        ExamTypeid = data['id']
        ExamTypeobj = ExamType.objects.filter(id=ExamTypeid,isActive=True,school_code=request.user.school_code).first()
        if ExamTypeobj is not None:
            ExamTypeexist = ExamType.objects.filter(TypeName=data['TypeName'],isActive= True,school_code=request.user.school_code).exclude(id=ExamTypeid).first()
            if ExamTypeexist is not None:
                return Response({"data":'',"response": {"n": 0, "msg": "ExamType already exist","status": "failure"}})
            else:
                serializer = ExamTypeSerializer(ExamTypeobj,data=data,partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({"data":serializer.data,"response": {"n": 1, "msg": "ExamType Updated successfully","status": "success"}})
                else:
                    return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't Update ExamType ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "ExamType not found ","status": "failure"}})


class deleteExamType(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        ExamTypeid = data['id']
        ExamTypeobj = ExamType.objects.filter(id=ExamTypeid,isActive=True,school_code=request.user.school_code).first()
        if ExamTypeobj is not None:
            data['isActive'] = False
            serializer = ExamTypeSerializer(ExamTypeobj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "ExamType Deleted successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't Delete ExamType ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "ExamType not found ","status": "failure"}})


class examscorelist(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        ExamTypeid = data['id']
        ExamTypeobj = ExamType.objects.filter(id=ExamTypeid,isActive=True,school_code=request.user.school_code)
        if ExamTypeobj.exists():
            exammarkser =  ExamTypeSerializer(ExamTypeobj,many=True)
            return Response({"data":exammarkser.data,"response": {"n": 1, "msg": "ExamType marks found successfully","status": "success"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "ExamType not found ","status": "failure"}})

class AddExamTypeMarks(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        data['isActive'] = True
        data['school_code']=request.user.school_code
        examtypeexist = ExamTypeMarks.objects.filter(Typeid=data['Typeid'],Marks=data['Marks'],isActive= True,school_code=request.user.school_code).first()
        if examtypeexist is not None:
            return Response({"data":'',"response": {"n": 0, "msg": "Type  with marks already exist","status": "failure"}})
        else:
            serializer = ExamTypeMarksSerializer1(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam Type Marks added successfully","status": "success"}})
            else:
                first_key, first_value = next(iter(serializer.errors.items()))
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": first_value[0],"status": "failure"}})

class exam_type_marks_list(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def get(self,request):
        ExamTypeobjs = ExamTypeMarks.objects.filter(isActive=True,school_code=request.user.school_code).order_by('-id')
        serializer = ExamTypeMarksSerializer(ExamTypeobjs,many=True)
        return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam Type Marks list found successfully","status": "success"}})
      
class delete_exam_type_marks(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        ExamTypeMarksid = data['id']
        ExamTypeMarksobj = ExamTypeMarks.objects.filter(id=ExamTypeMarksid,isActive=True,school_code=request.user.school_code).first()
        if ExamTypeMarksobj is not None:
            data['isActive'] = False
            serializer = ExamTypeMarksSerializer(ExamTypeMarksobj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "ExamTypeMarks Deleted successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't Delete ExamTypeMarks ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "ExamTypeMarks not found ","status": "failure"}})


class edit_exam_type_marks(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        ExamTypeMarksid = data['id']
        ExamTypeMarksobj = ExamTypeMarks.objects.filter(id=ExamTypeMarksid,isActive=True,school_code=request.user.school_code).first()
        if ExamTypeMarksobj is not None:
            ExamTypeMarksexist = ExamTypeMarks.objects.filter(Typeid=data['Typeid'],Marks=data['Marks'],isActive= True,school_code=request.user.school_code).exclude(id=ExamTypeMarksid).first()
            if ExamTypeMarksexist is not None:
                return Response({"data":'',"response": {"n": 0, "msg": "ExamTypeMarks already exist","status": "failure"}})
            else:
                serializer = ExamTypeMarksSerializer1(ExamTypeMarksobj,data=data,partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({"data":serializer.data,"response": {"n": 1, "msg": "ExamTypeMarks Updated successfully","status": "success"}})
                else:
                    return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't Update ExamTypeMarks ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "ExamTypeMarks not found ","status": "failure"}})

class get_marks_by_exam_type(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        ExamTypeid = data['id']
        if ExamTypeid is not None and ExamTypeid !='':
            
            ExamTypeobj = ExamType.objects.filter(id=ExamTypeid,isActive=True,school_code=request.user.school_code).first()
            if ExamTypeobj is not None:
                ExamTypMarkseobj = ExamTypeMarks.objects.filter(Typeid=ExamTypeobj.id,school_code=request.user.school_code)
                exammarkser =  ExamTypeMarksSerializer(ExamTypMarkseobj,many=True)
                return Response({"data":exammarkser.data,"response": {"n": 1, "msg": "Exam Type marks found successfully","status": "success"}})
            else:
                return Response({"data":[],"response": {"n": 0, "msg": "Exam Type not found ","status": "failure"}})

        else:
            return Response({"data":[],"response": {"n": 0, "msg": "Exam Type required","status": "failure"}})
            



def has_duplicate_value(classlist, key):
    seen = set()
    for d in classlist:
        value = d.get(key)
        if value in seen:
            return True
        seen.add(value)
    return False


class AddExam(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        schoolcode = request.user.school_code
        if request.POST.get('classlist') != "":
            classlist = json.loads(request.POST.get('classlist'))
        else:
            classlist = []
        print("daya",data)
        Date = data['date']
        Examstarttime = data['Examstarttime']
        Examendtime = data['Examendtime']
        SubjectId = data['SubjectId']
        ExamType = data['ExamType']
        totalMarks = data['totalMarks']
        passingmarks = data['passingmarks']
        reportTime = data['reportTime']
        Instructions = data['Instructions']
        AcademicYearId = int(data['AcademicYearId'])
        exam = int(data['exam'])
        exam_obj = Exam.objects.filter(id=exam,isActive=True,school_code=schoolcode).first()
        if exam_obj is None:
            return Response({"data":'',"response": {"n": 0, "msg": "exam  not found ","status": "failure"}})
        
        academic_obj = AcademicYear.objects.filter(id=AcademicYearId,Isdeleted=False,school_code=schoolcode).first()
        if academic_obj is None:
            return Response({"data":'',"response": {"n": 0, "msg": "Academic year not found ","status": "failure"}})
            
            

        for i in classlist :
            examobj = Exams.objects.filter(ClassId=i['ClassId'],Date=Date,Examstarttime__lt = Examendtime,Examendtime__gt=Examstarttime,school_code=schoolcode)
            if examobj.exists():
                return Response({"data":'',"response": {"n": 0, "msg": "Exam shedule already exists ! ","status": "failure"}})

        duplicateexist = has_duplicate_value(classlist, 'ClassId')

        if duplicateexist == False:
            for i in classlist:
                Exams.objects.create(ClassId_id=i['ClassId'],Date=Date,Examstarttime=Examstarttime,Examendtime=Examendtime,SubjectId_id=SubjectId,ExamType_id=ExamType,totalMarks=totalMarks,passingmarks=passingmarks,reportTime=reportTime,RoomNo=i['RoomNo'],InvigilatorId=i['InvigilatorId'],Instructions=Instructions,school_code=schoolcode,AcademicYearId=academic_obj,exam=exam_obj)
            return Response({"data":'',"response": {"n": 1, "msg": "Exams shedule created successfully","status": "success"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "Class repeated in list","status": "failed"}})
    


class Examlist(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def get(self,request):
        schoolcode = request.user.school_code
        Examobjs = Exams.objects.filter(isActive=True,school_code=schoolcode).order_by('-id')
        Examser = CustomExamsSerializer(Examobjs,many=True)
        for i in Examser.data:
            start_time = i['Examstarttime']
            end_time = i['Examendtime']

            # convert time string to datetime
            t1 = datetime.strptime(start_time, "%H:%M")

            t2 = datetime.strptime(end_time, "%H:%M")

            # get difference
            delta = t2 - t1

            i['totaltime'] = str(delta)+"hrs"
        return Response({"data":Examser.data,"response": {"n": 1, "msg": "Examslist found successfully","status": "success"}})
    

class Exambyid(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        id = request.data.get('id')
        Examobj = Exams.objects.filter(id=id,isActive=True).first()
        if Examobj is not None:
            serializer = ExamSerializer(Examobj)
            return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam found successfully","status": "success"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "Exam not found ","status": "failure"}})
        

class updateexam(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        id = request.data.get('id')
        data = request.data.copy()
        schoolcode = request.user.school_code
        Examobj = Exams.objects.filter(id=id,isActive=True).first()
        if Examobj is not None :
            reqdata={}
            reqdata['ClassId']  = data['ClassId']
            reqdata['Date'] = data['Date']
            reqdata['Examstarttime'] = data['Examstarttime']
            reqdata['Examendtime'] = data['Examendtime']
            reqdata['InvigilatorId'] = data['InvigilatorId']
            reqdata['SubjectId'] = data['SubjectId']
            reqdata['ExamType'] =  data['ExamType']
            reqdata['totalMarks'] = data['totalMarks']
            reqdata['reportTime'] =  data['reportTime']
            reqdata['RoomNo'] = data['RoomNo']
            reqdata['Instructions'] = data['Instructions']
            reqdata['exam'] = data['exam']
            reqdata['AcademicYearId'] = data['AcademicYearId']
            reqdata['passingmarks'] = data['passingmarks']

            serializer  = ExamSerializer(Examobj,data=reqdata,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam updated successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't update Exam ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "Exam not found ","status": "failure"}})
        

class deleteexam(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        Examid = data['id']
        Examobj = Exams.objects.filter(id=Examid,isActive=True).first()
        if Examobj is not None:
            data['isActive'] = False
            serializer = ExamSerializer(Examobj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam Deleted successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't Delete Exam ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "Exam not found ","status": "failure"}})
        


class exam_names_list(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def get(self,request):
        schoolcode = request.user.school_code
        Examobjs = Exam.objects.filter(isActive=True,school_code=schoolcode).order_by('Name')
        Examser = ExamNameSerializer(Examobjs,many=True)
        
        return Response({"data":Examser.data,"response": {"n": 1, "msg": "Exams found successfully","status": "success"}})
   
    
class academic_exam_list(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)

    def get(self,request):
        schoolcode = request.user.school_code
        Examobjs = Exams.objects.filter(isActive=True).order_by('exam_id','AcademicYearId_id',).distinct('exam_id','AcademicYearId_id',)
        serializer = CustomExamsSerializer1(Examobjs,many=True)
        
        newlist=serializer.data
        for i in newlist:
            print('sssss',i['exam_id'],i['AcademicYearId_id'])
            Exam_shedule_obj = Exams.objects.filter(exam=int(i['exam_id']),AcademicYearId=int(i['AcademicYearId_id']),isActive=True).order_by('Date')
            Exam_shedule_serializer = CustomExamsSerializer2(Exam_shedule_obj,many=True)
            i['shedule']=Exam_shedule_serializer.data
            for e in i['shedule']:
                start_time = e['Examstarttime']
                end_time = e['Examendtime']

                # convert time string to datetime
                t1 = datetime.strptime(start_time, "%H:%M")

                t2 = datetime.strptime(end_time, "%H:%M")

                # get difference
                delta = t2 - t1

                e['totaltime'] = str(delta)+"hrs"
            

        return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exams found successfully","status": "success"}})
    
class get_exam_timetable(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)

    def post(self,request):

        exam_id=request.POST.get('exam_id')
        AcademicYearId_id=request.POST.get('AcademicYearId')
        if exam_id is not None and exam_id != '':
            if AcademicYearId_id is not None and AcademicYearId_id != '':
                classid=request.POST.get('classid')
                school_code = request.user.school_code
                class_objs = Exams.objects.filter(isActive=True,exam_id=exam_id,AcademicYearId_id=AcademicYearId_id,school_code=school_code).order_by('ClassId_id').distinct('ClassId_id')
                if classid is not None and classid !='':
                    class_objs=class_objs.filter(ClassId=classid)
                if class_objs.exists() :   
                    serializer = CustomExamsSerializer1(class_objs,many=True)

                    newlist=serializer.data
                    for i in newlist:
                        Exam_shedule_obj = Exams.objects.filter(exam=int(i['exam_id']),AcademicYearId=int(i['AcademicYearId_id']),ClassId=int(i['ClassId_id']),isActive=True,school_code=school_code).order_by('Date')
                        Exam_shedule_serializer = CustomExamsSerializer2(Exam_shedule_obj,many=True)
                        i['shedule']=Exam_shedule_serializer.data
                        for e in i['shedule']:
                            start_time = e['Examstarttime']
                            end_time = e['Examendtime']

                            # convert time string to datetime
                            t1 = datetime.strptime(start_time, "%H:%M")

                            t2 = datetime.strptime(end_time, "%H:%M")

                            # get difference
                            delta = t2 - t1

                            e['totaltime'] = str(delta)+"hrs"
                    

                    return Response({"data":newlist,"response": {"n": 1, "msg": "Exams found successfully","status": "success"}})
                else:
                    return Response({"data":[],"response": {"n": 0, "msg": "No exam found","status": "failure"}})
            else:
                return Response({"data":[],"response": {"n": 0, "msg": "Please provide academic year id","status": "failure"}})
        else:
            return Response({"data":[],"response": {"n": 0, "msg": "Please provide exam id","status": "failure"}})


class uploadmarksheet(GenericAPIView):
    # authentication_classes=[userJWTAuthentication]
    # permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        datafile = request.FILES.get("excelfile")
        
        excel_file_path = datafile

        # Load the workbook
        workbook = load_workbook(excel_file_path)

        # Get all sheet names
        sheet_names = workbook.sheetnames

        # Dictionary to store data from all sheets
        all_sheet_data = {}

        # Loop through each sheet and extract data
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            all_sheet_data[sheet_name] = data

        # Now all_sheet_data contains data from all sheets in the Excel file
        # You can access each sheet's data using its name as key
        # for sheet_name, data in all_sheet_data.items():
        #     for row in data:
        #         print("row",row)
        
        return Response({"data":'',"response": {"n": 0, "msg": "Exam not found ","status": "failure"}})
    
    
    
 

class UploadExcelMarkSheet(GenericAPIView): 
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        dataset = Dataset()
        fileerrorlist=[]
        new_fees_distributions = request.FILES['file']

        AcademicYearId = request.POST.get('AcademicYearId')
        ClassId = request.POST.get('ClassId')
        Examid = request.POST.get('Exam')


        school_code=request.user.school_code
        if not new_fees_distributions.name.endswith('xlsx'):
            return Response({'data':[],"response":{"status":"failure",'msg': 'file format not supported','n':0}})
        
        imported_data = dataset.load(new_fees_distributions.read(), format='xlsx')

        if AcademicYearId is not None and AcademicYearId !='':
            academic_year_obj=AcademicYear.objects.filter(id=AcademicYearId,Isdeleted=False,school_code=school_code).first()
            if academic_year_obj is None:
                return Response({'data':[],"response":{"status":"failure",'msg': 'academic year not found','n':0}})
        else:
            return Response({'data':[],"response":{"status":"failure",'msg': 'Please select the academic year','n':0}})


        if ClassId is not None and ClassId !='':
            class_obj=Class.objects.filter(id=ClassId,isActive=True,school_code=school_code).first()
            if class_obj is None:
                return Response({'data':[],"response":{"status":"failure",'msg': 'class not found','n':0}})
        else:
            return Response({'data':[],"response":{"status":"failure",'msg': 'Please select the class','n':0}})

        if Examid is not None and Examid !='':
            Exam_obj=Exam.objects.filter(id=Examid,isActive=True,school_code=school_code).first()
            if Exam_obj is None:
                return Response({'data':[],"response":{"status":"failure",'msg': 'Exam not found','n':0}})
        else:
            return Response({'data':[],"response":{"status":"failure",'msg': 'Please select the Exam','n':0}})


        print("AcademicYearId",AcademicYearId,school_code)


        for i in imported_data:
            StudentCode=i[0]
            SubjectName = i[1]
            ObtainedMarks = i[2]
            PaperTypeName = i[3]

            if StudentCode is not None and StudentCode !="":
                StudentCode=''.join(StudentCode.split())
                students_obj=Students.objects.filter(StudentCode = str(StudentCode),school_code=school_code,isActive=True).first()
                if students_obj is not None:
                    student_serializer=StudentSerializer(students_obj)
                    students_class_obj=studentclassLog.objects.filter(classid = ClassId,AcademicyearId=AcademicYearId,studentId=student_serializer.data['id'],school_code=school_code,isActive=True).first()
                    if students_class_obj is not None:
                        if SubjectName is not None and SubjectName !="":
                            sub_obj = Subject.objects.filter(SubjectName__contains = str(SubjectName),isActive=True,school_code=school_code).first()
                            if sub_obj is not None:
                                if PaperTypeName is not None and PaperTypeName !='':
                                    paper_type_obj =ExamType.objects.filter(TypeName__contains=PaperTypeName,school_code=school_code,isActive=True).first()
                                    if paper_type_obj is not None:
                                        paper_type_serializer =ExamTypeSerializer(paper_type_obj)
                                        find_exam_obj=Exams.objects.filter(ClassId=ClassId,SubjectId=sub_obj.id,AcademicYearId=AcademicYearId,exam=Examid,school_code=school_code,ExamType=paper_type_serializer.data['id']).first()
                                        if find_exam_obj is not None:
                                            if ObtainedMarks is not None and ObtainedMarks !='':

                                                data={}
                                                data['AcademicYearId']=AcademicYearId
                                                data['ClassId']=ClassId

                                                data['RollNo']=students_class_obj.RollNo
                                                #exam name   
                                                print("data",find_exam_obj)
                                                if int(ObtainedMarks) < int(find_exam_obj.passingmarks):
                                                    data['Status']=0
                                                else:
                                                    data['Status']=1

                                                data['Exam']=Examid
                                                data['ObtainedMarks']=ObtainedMarks
                                                data['OutOfMarks']=find_exam_obj.totalMarks
                                                data['school_code']=school_code
                                                data['subID']=str(sub_obj.id)
                                                data['StudentId']=student_serializer.data['id']

                                                print("data",data)
                                                check_already_exist_obj=MarkSheet.objects.filter(AcademicYearId=data['AcademicYearId'],ClassId=data['ClassId'],Exam=data['Exam'],school_code=data['school_code'],subID=data['subID'],StudentId=data['StudentId']).first()
                                                if check_already_exist_obj is not None:
                                                    marksheet_serializer=MarkSheetSerializer(check_already_exist_obj,data=data,partial=True)
                                                else:
                                                    marksheet_serializer=MarkSheetSerializer(data=data)

                                                if marksheet_serializer.is_valid():
                                                    marksheet_serializer.save()
                                                else:

                                                    first_key, first_value = next(iter(marksheet_serializer.errors.items()))
                                                    reason = first_key + " : "+ first_value[0]
                                                    error = i + tuple([reason])
                                                    fileerrorlist.append(error)
                                                    continue  


                                            else:
                                                reason = 'please provide obtain marks'
                                                error = i + tuple([reason])
                                                fileerrorlist.append(error)
                                                continue  

                                        else:
                                            reason = 'exam not found'
                                            error = i + tuple([reason])
                                            fileerrorlist.append(error)
                                            continue  

                                    else:
                                        reason = 'Paper type with this '+PaperTypeName+' type not found'
                                        error = i + tuple([reason])
                                        fileerrorlist.append(error)
                                        continue  
                                else:
                                    reason = 'Paper type required'
                                    error = i + tuple([reason])
                                    fileerrorlist.append(error)
                                    continue  
                            else:
                                reason = 'subject with this subject name not found'
                                error = i + tuple([reason])
                                fileerrorlist.append(error)
                                continue  

                        else:
                            reason = 'please provide subject name'
                            error = i + tuple([reason])
                            fileerrorlist.append(error)
                            continue  
                    else:
                        reason = 'student with  this class  and academic  year an not found.'
                        error = i + tuple([reason])
                        fileerrorlist.append(error)
                        continue  
                else:
                    reason = 'student with this code '+StudentCode+' not found.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue  
            else:
                reason = 'student code is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue  


  



        if len(fileerrorlist) == 0:
            return Response({"data":'',"response": {"n": 1, "msg": "marsheet uploaded success fully","status": "success"}})
        else:
            return Response({"data":fileerrorlist,'headers':['Student Code','SubName','ObtainedMarks','Paper-Type','','','','Failure Reason'],"response": {"n": 2, "msg": "file has some issues","status": "failure"}})
    
    
    
    
class promotemarksheetexcel(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        Student = request.user.Student
        AcademicYearId = request.user.AcademicYearId
        promote_class = request.user.promote_class
        RollNo = request.user.RollNo
        SchoolCode = request.user.SchoolCode
        dataset = Dataset()
      
        new_product = request.FILES.get('classfile')

        if not new_product.name.endswith('xlsx'):
            return Response({"data":'',"response": {"n": 0, "msg": "Wrong File Format","status": "failure"}})

        imported_data = dataset.load(new_product.read(), format='xlsx')
        
        importDataList =[]
        notimporteddatalist = []
        for i in imported_data:
            if i[0] is not None:
                importDataList.append(i)
            else:
                notimporteddatalist.append(i)

        for i in importDataList:
            classexist = MarkSheet.objects.filter(ClassName__in=[i[0].lower(),i[0].upper()],Student=Student,AcademicYearId=AcademicYearId,promote_class=promote_class,RollNo=RollNo,SchoolCode=SchoolCode).first()
            if classexist is None:
                MarkSheet.objects.create(ClassName=i[0],Student=Student,AcademicYearId=AcademicYearId,promote_class=promote_class,RollNo=RollNo,SchoolCode=SchoolCode)

        return Response({"data":'done',"response": {"n": 1, "msg": "Promote Marksheet uploaded successfully","status": "success"}})

class GenerateMarkSheet(GenericAPIView): 
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        class_id = request.POST.get('class')
        academic_year_id = request.POST.get('yearid')
        exam_name_id = request.POST.get('exam_id')
        school_code=request.user.school_code
        student_ids_list = json.loads(request.POST.get('studentidlist'))
        marksheet_list=[]
        school_obj=School.objects.filter(school_code=school_code,isActive=True).first()
        school_serializer=schoolSerializer(school_obj)
        if class_id is not None and class_id !='':
            class_obj=Class.objects.filter(id=class_id,isActive=True,school_code=school_code).first()
            if class_obj is not None:
                if academic_year_id is not None and academic_year_id !='':
                    academic_year_obj=AcademicYear.objects.filter(id=academic_year_id,Isdeleted=False,school_code=school_code).first()
                    if academic_year_obj is not None:
                        if exam_name_id is not None and exam_name_id !='':
                            exam_name_obj = Exam.objects.filter(id=exam_name_id,isActive=True,school_code=school_code).first()
                            if exam_name_obj is not None:
                                if student_ids_list is not None and student_ids_list !='' and student_ids_list !=[]:
                                    for student_id in student_ids_list:
                                        marksheet={}
                                        marksheet['n']=1
                                        marksheet['exam_name']=str(exam_name_obj)
                                        student_obj = Students.objects.filter(id=student_id,isActive=True,school_code=school_code).first()
                                        if student_obj is not None:
                                            student_serializer=StudentSerializer2(student_obj)
                                            parent_obj=User.objects.filter(id=student_serializer.data['ParentId'],isActive=True,school_code=school_code).first()
                                            if parent_obj is not None:
                                                marksheet['ParentName']=str(parent_obj)
                                            else:
                                                marksheet['ParentName']=''

                                            marksheet['school_info']=school_serializer.data
                                            marksheet['student_info']=student_serializer.data
                                            marksheet['academic_year']=str(academic_year_obj)

                                            students_class_obj=studentclassLog.objects.filter(classid = class_id,AcademicyearId=academic_year_id,studentId=student_serializer.data['id'],school_code=school_code,isActive=True).first()
                                            if students_class_obj is not  None:
                                                student_class_serilaizer=custom_studentclassLogserializer(students_class_obj)
                                                marksheet['roll_no']=student_class_serilaizer.data['RollNo']
                                                marksheet['class_name']=student_class_serilaizer.data['classid']
                                                find_exams_subjects_obj=Exams.objects.filter(ClassId=class_id,AcademicYearId=academic_year_id,exam=exam_name_id,school_code=school_code)
                                                if find_exams_subjects_obj.exists():
                                                    exam_time_table_serializers=CustomExamsSerializer2(find_exams_subjects_obj,many=True)
                                                    marksheet['exam_time_table']=[]
                                                    marksheet['all_subject_total_marks']=0
                                                    marksheet['all_subject_total_marks_obtain']=0
                                                    Status='Pass'
                                                    for paper in exam_time_table_serializers.data:
                                                        papers={}
                                                        papers['totalMarks']=paper['totalMarks']
                                                        marksheet['all_subject_total_marks']+=int(paper['totalMarks'])
                                                        papers['SubjectId']=str(paper['SubjectId']).capitalize()
                                                        obtained_marks_obj=MarkSheet.objects.filter(AcademicYearId=academic_year_id,ClassId=student_class_serilaizer.data['classid_id'],StudentId=student_serializer.data['id'],
                                                                                                    subID=paper['SubjectId_id'],Exam=exam_name_id,school_code=school_code,isActive=True).first()
                                                        if obtained_marks_obj is not None:
                                                            marks_serializer=MarkSheetSerializer(obtained_marks_obj)
                                                            papers['ObtainedMarks']=marks_serializer.data['ObtainedMarks']
                                                            papers['Status']=marks_serializer.data['Status']
                                                            if papers['Status']=='0':
                                                                Status='Fail'
                                                            marksheet['all_subject_total_marks_obtain']+=int(marks_serializer.data['ObtainedMarks'])
                                                        else:
                                                            papers['ObtainedMarks']='NA'
                                                            papers['Status']='NA'

                                                        marksheet['exam_time_table'].append(papers)
                                                    marksheet['Percentage']=calculate_percentage(marksheet['all_subject_total_marks_obtain'],marksheet['all_subject_total_marks'])
                                                    marksheet['Status']=Status
                                                else:
                                                    marksheet['n']=0
                                                    marksheet['reason']='This exam is not present for the calss and academics'
                                            else:
                                                marksheet['n']=0
                                                marksheet['reason']='Student is not present in this class and academic year '
                                        else:
                                            marksheet['n']=0
                                            marksheet['reason']='Student is no longer present'
                                        marksheet_list.append(marksheet)     

                                    return Response({"data":{'marksheet_list':marksheet_list},"response": {"n": 1, "msg": "success","status": "success"}})

                                else:
                                    return Response({"data":{},"response": {"n": 0, "msg": "Please provide students id","status": "failure"}})
                            else:
                                return Response({"data":{},"response": {"n": 0, "msg": "exam name not found","status": "failure"}})
                        else:
                            return Response({"data":{},"response": {"n": 0, "msg": "Please provide students exam name","status": "failure"}})
                    else:
                        return Response({"data":{},"response": {"n": 0, "msg": " academic year not found","status": "failure"}})
                else:
                    return Response({"data":{},"response": {"n": 0, "msg": "Please provide students academic year","status": "failure"}})
            else:
                return Response({"data":{},"response": {"n": 0, "msg": "class not found","status": "failure"}})
        else:
            return Response({"data":{},"response": {"n": 0, "msg": "Please provide students class","status": "failure"}})

      
        


# class GenerateMarkSheet(GenericAPIView): 
#     authentication_classes=[userJWTAuthentication]
#     permission_classes = (permissions.IsAuthenticated,)
#     def post(self,request):
#         Status = ''
#         classid = request.POST.get('classid')
#         studentId = request.POST.get('studentId')
        
#         studenobj = Students.objects.filter(id=studentId).first()
#         studentname = studenobj.StudentName
#         ParentId = studenobj.ParentId
#         scode = studenobj.school_code
#         rollno = studenobj.RollNo
#         AcademicYearId = ''
#         StudentCode = ''
#         classobj = Class.objects.filter(id=classid).first()
#         classname = classobj.ClassName 
#         scname = studenobj.StudentCode
        
#         Username_obj = User.objects.filter(id=studenobj.ParentId).first()
#         parent_name = Username_obj.Username
                 
#         obj = MarkSheet.objects.filter(ClassId=classid,StudentId=studentId,isActive=True)
#         ser = MarkSheetSerializer(obj,many=True)
        
#         for s in ser.data:
            
#             if s['Status'] == '1':
#                 Status = 'PASS'
#             else:
#                 Status = 'FAIL'
                
#             studenobj = Students.objects.filter(id=s['Student']).first()
#             s['studentname'] = studenobj.StudentName
            
#             s['scode'] = s['SchoolCode']
#             s['rollno'] = s['RollNo']
#             classobj = Class.objects.filter(id=s['ClassId']).first()
#             s['classname'] = classobj.ClassName 
#             StudentCode = studenobj.StudentCode
            
#             subobj = Subject.objects.filter(id=s['SubID']).first()
#             s['subjname'] = subobj.SubjectName
            
#             examobj = Exam.objects.filter(id=s['Exam'],isActive=True,school_code=s['SchoolCode']).first()
#             s['exname'] = examobj.Name
            
#             AcademicYearId_obj = AcademicYear.objects.filter(id=s['AcademicYearId']).first()
#             AcademicYearId_name = str(AcademicYearId_obj.startdate) + '-' + str(AcademicYearId_obj.enddate)

            
#         return Response({"data":ser.data,"parent_name":parent_name,"AcademicYearId_name":AcademicYearId_name,"Status":Status,"studentname":studentname,"ParentId":ParentId,"scode":scode,"rollno":rollno,"classname":classname,"StudentCode":StudentCode,"response": {"n": 1, "msg": "Generate Marksheet successfully","status": "success"}})



# class multiplecandiateshortlist(GenericAPIView):
    # def post(self,request):
    #     studentId = json.loads(request.data.get('studentId'))
    #     if studentId is not None and studentId != "":
    #         for student_id  in studentId:
    #             studobject = studentclassLog.objects.filter(promote_class=False,id=student_id).first()
    #             if studobject is not None:
    #                 studobject.promote_class = True
    #                 studobject.save()
    #         response_={
    #             'status':1,
    #             'msg':'Student have been Promoted',
    #             'data':[]
    #         }
    #         return Response(response_,status=200) 
    #     response_={
    #         'status':0,
    #         'msg':'Student have not been Promoted',
    #         'data':[]
    #     }
    #     return Response(response_,status=200)
    
    
    
class MultipleStudentShortlist(GenericAPIView):
    def post(self, request):
        try:
            # Extract studentId from request data
            studentId = json.loads(request.data.get('checklist'))
            
            for i in studentId:
                studobject = studentclassLog.objects.filter(promote_class=False,id=i).first()
                if studobject is not None:
                    studobject.promote_class = True
                    studobject.save()
                    
                    class_obj = Class.objects.filter(ClassName=studobject.classid).first()
                    classname = list(class_obj.ClassName)
                    accturalclass_name = int(classname[0]) + 1
                    mainclass_name =  str(accturalclass_name) + str(classname[1])
                    classobj = Class.objects.filter(ClassName=mainclass_name).first()
                    
                    data = {
                        'promote_class':False,
                        'AcademicyearId':int(studobject.AcademicyearId.id),
                        'studentId':int(studobject.studentId.id),
                        'StudentCode':studobject.StudentCode,
                        'classid':int(classobj.id),
                        'RollNo':studobject.RollNo,
                        'school_code':studobject.school_code
                    }
                    studentclassser = studentclassLogserializer(data=data)
                    if studentclassser.is_valid():
                        studentclassser.save()
                    else:
                        print('studentclassser',studentclassser.errors)
                           
            response_ = {
                'status': 1,
                'msg': 'Students have been Promoted',
                'data': []
            }
            return Response(response_, status=200)
        
        except (ValueError, json.JSONDecodeError) as e:
            response_ = {
                'status': 0,
                'msg': f'Students have not been Promoted. Error: {str(e)}',
                'data': []
            }
            return Response(response_, status=400)
        
        
# class PromotedClassList(GenericAPIView):
#     def post(self, request):
#         studentclassLog.objects.filter()
class deleteexamname(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        Examid = data['id']
        Examobj = Exam.objects.filter(id=Examid,isActive=True,school_code=request.user.school_code).first()
        if Examobj is not None:
            data['isActive'] = False
            serializer = ExamSerializer(Examobj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam Deleted successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Couldn't Delete Exam ! ","status": "failure"}})
        else:
            return Response({"data":'',"response": {"n": 0, "msg": "Exam not found ","status": "failure"}})


class add_examname(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        data = request.data.copy()
        data['isActive'] = True
        data['school_code']=request.user.school_code
        exam_exist = Exam.objects.filter(Name=data['Name'],isActive= True,school_code=data['school_code']).first()
        if exam_exist is not None:
            return Response({"data":'',"response": {"n": 0, "msg": "Exam Name already exist","status": "failure"}})
        else:
            serializer = ExamNameSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data,"response": {"n": 1, "msg": "Exam Name added successfully","status": "success"}})
            else:
                return Response({"data":serializer.errors,"response": {"n": 0, "msg": "Exam Name not added ","status": "failure"}})


class examtypebyexcel(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        
        school_code = request.user.school_code
        dataset = Dataset()
        fileerrorlist=[]
        new_product = request.FILES.get('classfile')

        if not new_product.name.endswith('xlsx'):
            return Response({'data':[],"response":{"status":"failure",'msg': 'file format not supported','n':0}})

        imported_data = dataset.load(new_product.read(), format='xlsx')
        for i in imported_data:
            TypeName =i[0]
            data={}
            
            if TypeName is not None and TypeName !="":
                data['TypeName']= TypeName
                ExamType_exist = ExamType.objects.filter(TypeName__in = [data['TypeName'].strip().capitalize(),data['TypeName'].strip(),data['TypeName'].title(),data['TypeName'].upper(),data['TypeName'].lower(),data['TypeName']],isActive= True,school_code=school_code).first()
                if ExamType_exist is None:
                    ExamType.objects.create(TypeName= data['TypeName'],school_code=school_code)
                else:
                    reason = 'Exam type name already exits.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Exam type name is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
  

        if len(fileerrorlist) == 0:
            return Response({"data":'done',"response": {"n": 1, "msg": "Exam Type Name uploaded successfully","status": "success"}})
        else:
            return Response({"data":fileerrorlist,'headers':['TypeName','Failure Reason'],"response": {"n": 2, "msg": "file has some issues","status": "failure"}})


class examnamedatabyexcel(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        school_code = request.user.school_code
        dataset = Dataset()
        fileerrorlist=[]
        new_product = request.FILES.get('classfile')

        if not new_product.name.endswith('xlsx'):
            return Response({'data':[],"response":{"status":"failure",'msg': 'file format not supported','n':0}})

        imported_data = dataset.load(new_product.read(), format='xlsx')
        for i in imported_data:
            Name =i[0]
            data={}
            
            if Name is not None and Name !="":
                data['Name']= Name
                Examname_exist = Exam.objects.filter(Name__in = [data['Name'].strip().capitalize(),data['Name'].strip(),data['Name'].title(),data['Name'].upper(),data['Name'].lower(),data['Name']],isActive= True,school_code=school_code).first()
                if Examname_exist is None:
                    Exam.objects.create(Name= data['Name'],school_code=school_code)
                else:
                    reason = 'Exam name already exits.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Exam name is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
  
        if len(fileerrorlist) == 0:
            return Response({"data":'done',"response": {"n": 1, "msg": "Exam Name uploaded successfully","status": "success"}})
        else:
            return Response({"data":fileerrorlist,'headers':['ExamName','Failure Reason'],"response": {"n": 2, "msg": "file has some issues","status": "failure"}})



class examscheduldatabyexcel(GenericAPIView):
    authentication_classes=[userJWTAuthentication]
    permission_classes = (permissions.IsAuthenticated,)
    def post(self,request):
        school_code = request.user.school_code
        dataset = Dataset()
        fileerrorlist=[]
        new_product = request.FILES.get('classfile')

        if not new_product.name.endswith('xlsx'):
            return Response({'data':[],"response":{"status":"failure",'msg': 'file format not supported','n':0}})

        imported_data = dataset.load(new_product.read(), format='xlsx')
        for i in imported_data:
            
            ClassId =i[0]
            Date =i[1]
            Examstarttime =i[2]
            Examendtime =i[3]
            InvigilatorId =i[4]
            SubjectId =i[5]
            Examtype =i[6]
            totalMarks =i[7]
            reportTime =i[8]
            RoomNo =i[9]
            Instructions =i[10]
            AcademicYearId =i[11]
            exam = i[11]
            data={}
            
            if ClassId is not None and ClassId !="":
                ClassName_exist = Class.objects.filter(ClassName=i[0],isActive= True,school_code=school_code).first()
                if ClassName_exist is not None:
                    data['ClassId']= ClassName_exist.id
                else:
                    reason = 'Class name not found.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Class is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
                
            data['Date']= Date
            data['Examstarttime']= Examstarttime
            data['Examendtime']= Examendtime
            
            if InvigilatorId is not None and InvigilatorId !="":
                data['InvigilatorId']= InvigilatorId
                obj = User.objects.filter(role_id=4,isActive=True).first()
                if obj is not None:
                    data['InvigilatorId']= obj.id
                    
                else:
                    reason = 'InvigilatorId not found.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'InvigilatorId is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
                
            
            if SubjectId is not None and SubjectId !="":
                data['SubjectId']= SubjectId
                print('hii',data['SubjectId'])
                subject_exist = Subject.objects.filter(SubjectName__in=[i[5].strip().capitalize(),i[5].strip(),i[5].title(),i[5].upper(),i[5].lower(),i[5]],isActive= True,school_code=school_code).first()
                print('subject_exist',subject_exist)
                if subject_exist is not None:
                    data['SubjectId']= subject_exist.id
                else:
                    reason = 'Subject not found.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Subject is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
            
            if Examtype is not None and Examtype !="":
                data['Examtype']= Examtype
                print('i6',i[6])
                examtype_exist = ExamType.objects.filter(TypeName__in=[i[6].strip().capitalize(),i[6].strip(),i[6].title(),i[6].upper(),i[6].lower(),i[6]],isActive= True,school_code=school_code).first()
                print('examtype_exist',examtype_exist)
                if examtype_exist is not None:
                    data['Examtype']= examtype_exist.id
                else:
                    reason = 'Exam type name already exits.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Exam type name is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
            
            
            data['totalMarks']= totalMarks
            data['reportTime']= reportTime
            data['RoomNo']= RoomNo
            data['Instructions']= Instructions
            
            if AcademicYearId is not None and AcademicYearId !="":
                data['AcademicYearId']= AcademicYearId 
                AcademicYear_exist = AcademicYear.objects.filter(startdate__in=[i[11].split('to')[0].strip()],enddate__in=[i[11].split('to')[1].strip()],school_code=school_code).first()
                if AcademicYear_exist is not None:
                    data['AcademicYearId']= AcademicYear_exist.id
                else:
                    reason = 'Academic Year not found.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Academic Year is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue 
            
            if exam is not None and exam !="":
                data['exam']= exam
                Examname_exist = Exam.objects.filter(Name__in=[i[12].strip().capitalize(),i[12].strip(),i[12].title(),i[12].upper(),i[12].lower(),i[12]],isActive= True,school_code=school_code).first()
                if Examname_exist is not None:
                    data['exam']= Examname_exist.id
                else:
                    reason = 'Exam name already exits.'
                    error = i + tuple([reason])
                    fileerrorlist.append(error)
                    continue 
            else:
                reason = 'Exam name is required.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue
            
            
            Date =i[1]
            Exam_schedule = Exams.objects.filter(reportTime=data['reportTime'],ClassId_id=data['ClassId'],Date=data['Date'],Examstarttime=data['Examstarttime'],Examendtime=data['Examendtime'],InvigilatorId=data['InvigilatorId'],totalMarks=data['totalMarks'],AcademicYearId_id=data['AcademicYearId'],SubjectId_id=data['SubjectId'],ExamType_id=data['Examtype'],Instructions=data['Instructions'],RoomNo=data['RoomNo'],exam_id=data['exam'],isActive= True,school_code=school_code).first()
            if Exam_schedule is None:
                Exams.objects.create(reportTime=data['reportTime'],ClassId_id=data['ClassId'],Date=data['Date'],Examstarttime=data['Examstarttime'],Examendtime=data['Examendtime'],InvigilatorId=data['InvigilatorId'],totalMarks=data['totalMarks'],AcademicYearId_id=data['AcademicYearId'],SubjectId_id=data['SubjectId'],ExamType_id=data['Examtype'],Instructions=data['Instructions'],RoomNo=data['RoomNo'],exam_id=data['exam'],school_code=school_code)
            else:
                reason = 'Exam schedule already exits.'
                error = i + tuple([reason])
                fileerrorlist.append(error)
                continue 
        # else:
        #     reason = 'Exam schedule is required.'
        #     error = i + tuple([reason])
        #     fileerrorlist.append(error)
        #     continue
  

        if len(fileerrorlist) == 0:
            return Response({"data":'done',"response": {"n": 1, "msg": "Exam schedule uploaded successfully","status": "success"}})
        else:
            return Response({"data":fileerrorlist,'headers':['Class','Date','ExamStartTime','ExamEndTime','InvigilatorId','Subject','ExamType','TotalMarks','ReportTime','RoomNo','Instructions','AcademicYear','ExamName','Failure Reason'],"response": {"n": 2, "msg": "file has some issues","status": "failure"}})





















